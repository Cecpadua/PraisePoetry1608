import openpyxl
from pptx import Presentation
from PIL import Image, ImageEnhance
import pytesseract
from io import BytesIO
import cv2
import numpy as np


def save_to_excel(filename, comments, initials):
    # 创建或加载现有的Excel文件
    excel_file_path = "歌词.xlsx"
    if os.path.exists(excel_file_path):
        workbook = openpyxl.load_workbook(excel_file_path)
    else:
        workbook = openpyxl.Workbook()

    # 选择默认的活动工作表
    sheet = workbook.active

    # 如果是新的Excel文件，添加标题行
    if not os.path.exists(excel_file_path) or sheet.cell(row=1, column=1).value is None:
        sheet['A1'] = '文件名'
        sheet['B1'] = '歌词'
        sheet['C1'] = '拼音首字母'

    # 在下一行添加数据
    next_row = sheet.max_row + 1
    sheet.cell(row=next_row, column=1, value=filename)
    sheet.cell(row=next_row, column=2, value=comments)
    sheet.cell(row=next_row, column=3, value=initials)

    # 保存Excel文件
    workbook.save(excel_file_path)

    print(f'数据已保存到Excel文件: {excel_file_path}')
    

        
import os
    
for file in os.listdir("D:/Users/yihao zhuo/Desktop/ppt/dir/ppt 16-9"):
    if file.endswith(".pptx"):
        pptx_path = f"D:/Users/yihao zhuo/Desktop/ppt/dir/ppt 16-9/{file}"
        presentation = Presentation(pptx_path)
        save_to_excel(file, presentation.core_properties.comments, presentation.core_properties.subject)
